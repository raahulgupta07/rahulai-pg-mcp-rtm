#!/usr/bin/env python3
"""
Job Manager - Job ID Generation and Job Tracking
"""

import os
from datetime import datetime
from typing import Optional, Dict
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.database import get_database

# ── Excel Style Constants ──
DARK_BG = PatternFill(start_color="383832", end_color="383832", fill_type="solid")
DARK_FONT = Font(color="FEFFD6", bold=True, size=10, name="Calibri")
GREEN_BG = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
AMBER_BG = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
RED_BG = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
TEAL_BG = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
ALT_ROW = PatternFill(start_color="FCF9EF", end_color="FCF9EF", fill_type="solid")
WHITE_ROW = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='383832'),
    right=Side(style='thin', color='383832'),
    top=Side(style='medium', color='383832'),
    bottom=Side(style='medium', color='383832'),
)


def _style_sheet(ws, df, title=None):
    """Apply professional styling to a worksheet."""
    start_row = 1

    # Title row
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(color="FEFFD6", bold=True, size=14, name="Calibri")
        title_cell.fill = DARK_BG
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        # Fill rest of title row with dark bg
        for col_idx in range(2, len(df.columns) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = DARK_BG
        start_row = 2

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = DARK_FONT
        cell.fill = DARK_BG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    ws.row_dimensions[start_row].height = 22

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=start_row + 1 + row_idx, column=col_idx)

            # Handle value types
            if pd.isna(value):
                cell.value = ""
            elif isinstance(value, (pd.Timestamp, datetime)):
                cell.value = value
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, float):
                cell.value = value
                col_str = str(col_name)
                if 'Pct' in col_str or 'Contribution' in col_str or '%' in col_str:
                    cell.number_format = '0.00%' if value <= 1 else '0.00'
                else:
                    cell.number_format = '#,##0'
            elif isinstance(value, (int, np.integer)):
                cell.value = int(value)
                cell.number_format = '#,##0'
            else:
                cell.value = str(value)

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = WHITE_ROW
            else:
                cell.fill = ALT_ROW

            cell.border = THIN_BORDER
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(vertical="center")

            # Color-code Classification column
            if col_name == "Classification":
                val = str(value)
                if "Class A" in val and "Local" in val:
                    cell.fill = TEAL_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="006F7C")
                elif "Class A" in val:
                    cell.fill = GREEN_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="007518")
                elif "Class B" in val:
                    cell.fill = AMBER_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="856404")
                elif "Class C" in val:
                    cell.fill = RED_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="BE2D06")

            # Color-code Visit_Frequency
            if col_name == "Visit_Frequency":
                if str(value) == "F4":
                    cell.fill = GREEN_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="007518")
                elif str(value) == "F2":
                    cell.fill = AMBER_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="856404")

            # Color-code AI_Growth_Signal
            if col_name == "AI_Growth_Signal":
                val = str(value).lower()
                if "grow" in val:
                    cell.fill = GREEN_BG
                    cell.font = Font(size=10, name="Calibri", color="007518")
                elif "declin" in val:
                    cell.fill = RED_BG
                    cell.font = Font(size=10, name="Calibri", color="BE2D06")

            # Color-code AI_Risk_Level
            if col_name == "AI_Risk_Level":
                val = str(value).lower()
                if "high" in val:
                    cell.fill = RED_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="BE2D06")
                elif "medium" in val:
                    cell.fill = AMBER_BG
                    cell.font = Font(size=10, name="Calibri", color="856404")
                elif "low" in val:
                    cell.fill = GREEN_BG
                    cell.font = Font(size=10, name="Calibri", color="007518")

            # Color-code Workload_Status
            if col_name == "Workload_Status":
                if str(value) == "OK":
                    cell.fill = GREEN_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="007518")
                elif str(value) == "BELOW_MIN":
                    cell.fill = RED_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="BE2D06")
                elif str(value) == "ABOVE_MAX":
                    cell.fill = AMBER_BG
                    cell.font = Font(size=10, name="Calibri", bold=True, color="856404")

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(df.columns[col_idx - 1])) + 2
        for row_idx in range(min(50, len(df))):
            cell_val = ws.cell(row=start_row + 1 + row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)) + 2, 40))
        ws.column_dimensions[col_letter].width = max(max_len, 8)

    # Auto-filter
    if len(df) > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"


class JobManager:
    """Manages job creation, tracking, and execution"""

    def __init__(self):
        self.db = get_database()
        self.current_job_id = None
        self.outputs_dir = self._get_outputs_dir()

    def _get_outputs_dir(self) -> str:
        """Get outputs directory path"""
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        outputs_dir = os.path.join(base_dir, "outputs")
        if not os.path.exists(outputs_dir):
            os.makedirs(outputs_dir)
        return outputs_dir

    def generate_job_id(self) -> str:
        """
        Generate unique job ID
        Format: RTM-YYYYMMDD-XXXX
        """
        today = datetime.now().strftime("%Y%m%d")

        # Get all jobs from today
        all_jobs = self.db.get_all_jobs(limit=100)
        today_jobs = [j for j in all_jobs if j["job_id"].startswith(f"RTM-{today}")]

        if today_jobs:
            # Get the last sequence number
            sequences = [int(j["job_id"].split("-")[-1]) for j in today_jobs]
            next_seq = max(sequences) + 1
        else:
            next_seq = 1

        job_id = f"RTM-{today}-{next_seq:04d}"
        return job_id

    def start_job(
        self,
        sales_filename: str,
        customer_filename: str = "",
        item_filename: str = "",
        threshold_a: int = 80,
        threshold_b: int = 95,
    ) -> str:
        """
        Start a new job. Retries on a job-id collision so two classifications
        running at the same time never clash.
        """
        last_err = None
        for _ in range(12):
            job_id = self.generate_job_id()
            try:
                self.db.create_job(
                    job_id,
                    sales_filename,
                    customer_filename,
                    item_filename,
                    threshold_a,
                    threshold_b,
                )
                self.current_job_id = job_id
                return job_id
            except Exception as e:  # PK collision under concurrency → retry
                last_err = e
        raise RuntimeError(f"Could not allocate a job id: {last_err}")

    def create_excel_report(self, results_df: pd.DataFrame, workload_df: pd.DataFrame = None,
                            meta: dict = None, comparison: dict = None) -> bytes:
        """Create multi-sheet Excel report with professional styling, per-branch
        breakdown, a run-info stamp, and a vs-previous-run comparison."""
        buf = io.BytesIO()
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # ── Sheet 1: All Results ──
        ws = wb.create_sheet("All Results")
        _style_sheet(ws, results_df, "RTM AGENT \u2014 ALL OUTLET RESULTS")

        # ── Sheet 2: Overall Summary ──
        cus_col = "Cus.Code" if "Cus.Code" in results_df.columns else "Cus_Code"
        summary = (
            results_df.groupby("Classification")
            .agg({cus_col: "count", "TotalSales_2Yr": ["sum", "mean"]})
            .round(2)
        )
        summary.columns = ["Count", "TotalSales", "AvgSales"]
        summary["SalesPct"] = (
            summary["TotalSales"] / summary["TotalSales"].sum() * 100
        ).round(2)
        summary_df = summary.reset_index()
        ws = wb.create_sheet("Overall Summary")
        _style_sheet(ws, summary_df, "CLASSIFICATION SUMMARY")

        # ── Sheet 3: Branch Summary (matrix: Branch x Class) ──
        if "BranchName" in results_df.columns:
            branch_class = (
                results_df.groupby(["BranchName", "Classification"])
                .agg({cus_col: "count", "TotalSales_2Yr": "sum"})
                .round(0)
            )
            branch_class.columns = ["Outlets", "Revenue"]
            branch_class = branch_class.reset_index()

            branch_totals = results_df.groupby("BranchName").agg(
                Total_Outlets=(cus_col, "count"),
                Total_Revenue=("TotalSales_2Yr", "sum"),
            ).reset_index()

            pivot_counts = branch_class.pivot_table(
                index="BranchName", columns="Classification",
                values="Outlets", fill_value=0, aggfunc="sum",
            )
            pivot_revenue = branch_class.pivot_table(
                index="BranchName", columns="Classification",
                values="Revenue", fill_value=0, aggfunc="sum",
            )

            branch_summary = branch_totals.copy()
            for cls in ["Class A", "Class B", "Class C", "Class A Local (F4)"]:
                if cls in pivot_counts.columns:
                    branch_summary[f"{cls}_Count"] = branch_summary["BranchName"].map(
                        pivot_counts[cls]
                    ).fillna(0).astype(int)
                    branch_summary[f"{cls}_Revenue"] = branch_summary["BranchName"].map(
                        pivot_revenue[cls]
                    ).fillna(0).astype(int)
                    branch_summary[f"{cls}_Pct"] = (
                        branch_summary[f"{cls}_Revenue"] / branch_summary["Total_Revenue"] * 100
                    ).round(1)

            branch_summary = branch_summary.sort_values("Total_Revenue", ascending=False)
            ws = wb.create_sheet("Branch Summary")
            _style_sheet(ws, branch_summary, "BRANCH PERFORMANCE MATRIX")

        # ── Sheets 4-N: One sheet per branch ──
        if "BranchName" in results_df.columns:
            branches = sorted(results_df["BranchName"].unique())
            for branch in branches:
                branch_df = results_df[results_df["BranchName"] == branch].copy()
                branch_df = branch_df.sort_values("TotalSales_2Yr", ascending=False)
                sheet_name = str(branch)[:31]
                ws = wb.create_sheet(sheet_name)
                _style_sheet(ws, branch_df, f"{str(branch).upper()} \u2014 OUTLET CLASSIFICATION")

        # ── Top 50 (across all branches) ──
        top_cols = ["BranchName", "Cus.Code", "Cus.Name", "Classification", "TotalSales_2Yr"]
        available_top = [c for c in top_cols if c in results_df.columns]
        if "TotalSales_12M" in results_df.columns:
            available_top.append("TotalSales_12M")
        top_customers = results_df.nlargest(50, "TotalSales_2Yr")[available_top]
        ws = wb.create_sheet("Top 50")
        _style_sheet(ws, top_customers, "TOP 50 OUTLETS BY REVENUE")

        # ── AI Action Plan ──
        ai_cols = [
            "BranchName", "Cus.Code", "Cus.Name", "Classification",
            "TotalSales_2Yr", "AI_Growth_Signal", "AI_Risk_Level",
            "AI_Visit_Priority", "AI_Action", "AI_Insight",
        ]
        available_ai = [c for c in ai_cols if c in results_df.columns]
        if len(available_ai) > 4:
            ai_df = results_df[available_ai].copy()
            ws = wb.create_sheet("AI Action Plan")
            _style_sheet(ws, ai_df, "AI ENRICHMENT \u2014 ACTION PLAN")

        # ── Seller Workload ──
        if workload_df is not None and not workload_df.empty:
            ws = wb.create_sheet("Seller Workload")
            _style_sheet(ws, workload_df, "SELLER WORKLOAD ANALYSIS")

        # ── vs Previous Run comparison ──
        if comparison and comparison.get("has_previous"):
            cmp_cols = ["Metric", "This Run", "Previous Run", "Change", "Change %", "Remark"]

            def cmp_df(rows):
                return pd.DataFrame([{
                    "Metric": r["metric"],
                    "This Run": r["current"],
                    "Previous Run": r["previous"],
                    "Change": r["change"],
                    "Change %": r["pct"],
                    "Remark": r["remark"],
                } for r in rows], columns=cmp_cols)

            summ = list(comparison.get("summary", []))
            mv = comparison.get("movement", {})
            for label, key in [
                ("Outlets Upgraded", "upgraded"), ("Outlets Downgraded", "downgraded"),
                ("Outlets Unchanged", "unchanged"), ("New Outlets", "new"),
                ("Lost Outlets", "lost"),
            ]:
                summ.append({"metric": label, "current": mv.get(key, 0),
                             "previous": "", "change": "", "pct": "", "remark": ""})
            prev_id = comparison.get("previous", {}).get("job_id", "")
            ws = wb.create_sheet("Run Comparison")
            _style_sheet(ws, cmp_df(summ), f"THIS RUN vs PREVIOUS — {prev_id}")

            if comparison.get("by_channel"):
                ws = wb.create_sheet("Compare Channels")
                _style_sheet(ws, cmp_df(comparison["by_channel"]), "OUTLETS BY CHANNEL — vs PREVIOUS RUN")
            if comparison.get("by_branch"):
                ws = wb.create_sheet("Compare Branches")
                _style_sheet(ws, cmp_df(comparison["by_branch"]), "OUTLETS BY BRANCH — vs PREVIOUS RUN")

        # ── Run Info (placed first) ──
        if meta:
            info_df = pd.DataFrame([
                {"Field": "Job ID", "Value": str(meta.get("job_id", ""))},
                {"Field": "Run date", "Value": str(meta.get("run_date", ""))},
                {"Field": "Run by", "Value": str(meta.get("run_by", ""))},
                {"Field": "Run by — user ID", "Value": str(meta.get("run_by_id", ""))},
                {"Field": "Rule config version", "Value": str(meta.get("rule_version", ""))},
                {"Field": "Rules last modified", "Value": str(meta.get("rule_modified_at", ""))},
                {"Field": "Rules modified by", "Value": str(meta.get("rule_modified_by", ""))},
                {"Field": "LLM tokens used", "Value": str(meta.get("llm_tokens", ""))},
                {"Field": "LLM cost (USD)", "Value": str(meta.get("llm_cost", ""))},
            ], columns=["Field", "Value"])
            ws = wb.create_sheet("Run Info", 0)
            _style_sheet(ws, info_df, "RUN INFORMATION")

        wb.save(buf)
        return buf.getvalue()

    def complete_job(
        self,
        job_id: str,
        results_df: pd.DataFrame,
        date_from: str = None,
        date_to: str = None,
        result_excel: bytes = None,
        workload_df: pd.DataFrame = None,
    ) -> bool:
        """Mark a job completed and persist its results. job_id is explicit so
        concurrent classifications never clobber each other."""
        if not job_id:
            raise ValueError("No job_id given to complete_job")

        self.db.update_job_results(job_id, results_df, date_from, date_to)

        if result_excel is None:
            result_excel = self.create_excel_report(results_df, workload_df=workload_df)

        excel_path = os.path.join(self.outputs_dir, f"{job_id}.xlsx")
        with open(excel_path, "wb") as f:
            f.write(result_excel)

        self.db.set_job_result_path(job_id, excel_path)
        return True

    def fail_job(self, job_id: str, error_message: str) -> bool:
        """Mark a job failed (job_id explicit — concurrency-safe)."""
        if not job_id:
            return False
        self.db.update_job_status(job_id, "failed", error_message)
        return True

    def get_current_job_id(self) -> Optional[str]:
        """Get current job ID"""
        return self.current_job_id

    def get_job(self, job_id: str) -> Optional[Dict]:
        """Get job details"""
        return self.db.get_job(job_id)

    def get_all_jobs(self, limit: int = 50) -> list:
        """Get all jobs"""
        return self.db.get_all_jobs(limit)

    def get_job_results(self, job_id: str) -> pd.DataFrame:
        """Get job results"""
        return self.db.get_job_results(job_id)

    def get_excel_path(self, job_id: str) -> Optional[str]:
        """Get Excel file path for a job"""
        job = self.db.get_job(job_id)
        if job and job.get("result_path"):
            return job["result_path"]
        return None

    def read_excel_file(self, job_id: str) -> Optional[bytes]:
        """Read Excel file for a job"""
        excel_path = self.get_excel_path(job_id)
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as f:
                return f.read()
        return None


# Singleton instance
_job_manager = None


def get_job_manager() -> JobManager:
    """Get job manager singleton"""
    global _job_manager
    if _job_manager is None:
        _job_manager = JobManager()
    return _job_manager
